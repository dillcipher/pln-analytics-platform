"""Memory-safe, restart-safe DLPD XLSX ingestion.

Large DLPD workbooks must never be materialised as a complete pandas
DataFrame. This module keeps the existing MonthlyMerger for small/non-DLPD
datasets and replaces only DLPD processing with a bounded
openpyxl -> pandas -> parquet pipeline.

Coordinate enrichment is disk-backed: the old implementation loaded the
entire CUSTOMER_LOCATION parquet for every DLPD chunk. That could push the
500 MB production container over its memory limit. We now build a local
SQLite IDPEL -> coordinate index once per monthly master and join each DLPD
chunk against that index.
"""

from __future__ import annotations

import logging
import os
import shutil
import sqlite3
import uuid
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.etl.merger.monthly_merger import MonthlyMerger
from app.etl.transformers.dlpd_transformer import DLPDTransformer
from app.etl.validator.validator import DatasetValidator

logger = logging.getLogger(__name__)

_INSTALLED = False
_ORIGINAL_MERGE = MonthlyMerger.merge

# Conservative default for the 500 MB production runtime.
_CHUNK_ROWS = max(
    500,
    int(os.getenv("DLPD_STREAM_CHUNK_ROWS", "2000")),
)

_COMPLETED_RUNS: dict[tuple[str, tuple[str, ...]], dict[str, Path]] = {}


def _source_key(dataset: str, files: list[Path]) -> tuple[str, tuple[str, ...]]:
    return (
        dataset,
        tuple(sorted(str(Path(path).resolve()) for path in files)),
    )


def _iter_excel_chunks(filepath: Path, dataset: str):
    """Yield bounded pandas frames from an XLSX worksheet."""
    filepath = Path(filepath)
    if filepath.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError(
            f"Streaming DLPD requires XLSX/XLSM format: {filepath.name}"
        )

    sheet = DatasetValidator.get_sheet_name(filepath, dataset)
    header = DatasetValidator.detect_header_row(filepath, sheet)
    workbook = load_workbook(filepath, read_only=True, data_only=True)

    try:
        worksheet = workbook[sheet]
        header_values = next(
            worksheet.iter_rows(
                min_row=header + 1,
                max_row=header + 1,
                values_only=True,
            ),
            None,
        )
        if not header_values:
            return

        columns = [
            str(value).strip().upper() if value is not None else ""
            for value in header_values
        ]
        if not any(columns):
            return

        batch: list[tuple] = []
        for row in worksheet.iter_rows(
            min_row=header + 2,
            values_only=True,
        ):
            batch.append(tuple(row[: len(columns)]))
            if len(batch) >= _CHUNK_ROWS:
                yield pd.DataFrame.from_records(batch, columns=columns)
                batch.clear()

        if batch:
            yield pd.DataFrame.from_records(batch, columns=columns)
    finally:
        workbook.close()


def _normalise_output_frame(frame: pd.DataFrame) -> pd.DataFrame:
    frame = frame.copy()
    frame.columns = frame.columns.map(str)
    for column in frame.columns:
        if pd.api.types.is_object_dtype(frame[column]):
            frame[column] = (
                frame[column].fillna("").astype(str).str.strip()
            )
    return frame


def _deduplicate_first_idpel(
    frame: pd.DataFrame,
    seen_db: sqlite3.Connection,
) -> pd.DataFrame:
    """Replicate first-IDPEL-wins without a Python set of all customers."""
    if frame.empty or "IDPEL" not in frame.columns:
        return frame

    ids = frame["IDPEL"].fillna("").astype(str).str.strip()
    frame = frame.loc[ids.ne("")].copy()
    if frame.empty:
        return frame

    ids = frame["IDPEL"].astype(str).str.strip()
    frame = frame.loc[~ids.duplicated(keep="first")].copy()
    ids = frame["IDPEL"].astype(str).str.strip()

    seen_db.execute("DELETE FROM batch_ids")
    seen_db.executemany(
        "INSERT OR IGNORE INTO batch_ids(idpel) VALUES (?)",
        ((value,) for value in ids.tolist()),
    )

    new_ids = {
        row[0]
        for row in seen_db.execute(
            """
            SELECT b.idpel
            FROM batch_ids AS b
            LEFT JOIN seen_idpel AS s ON s.idpel = b.idpel
            WHERE s.idpel IS NULL
            """
        )
    }
    if not new_ids:
        return frame.iloc[0:0].copy()

    seen_db.executemany(
        "INSERT INTO seen_idpel(idpel) VALUES (?)",
        ((value,) for value in new_ids),
    )
    seen_db.commit()
    return frame.loc[ids.isin(new_ids)].copy()


class _CoordinateIndex:
    """Disk-backed coordinate lookup shared by every DLPD chunk in a run."""

    def __init__(self, db_path: Path, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.conn = sqlite3.connect(db_path)
        self.conn.execute(
            """
            CREATE TABLE coordinate_index (
                month TEXT NOT NULL,
                idpel TEXT NOT NULL,
                koordinat_x REAL,
                koordinat_y REAL,
                PRIMARY KEY (month, idpel)
            )
            """
        )
        self.conn.execute(
            "CREATE INDEX coordinate_index_idpel ON coordinate_index(idpel)"
        )
        self.conn.execute(
            "CREATE TEMP TABLE lookup_ids (idpel TEXT PRIMARY KEY)"
        )
        self.conn.execute(
            "CREATE TABLE loaded_months (month TEXT PRIMARY KEY, available INTEGER NOT NULL)"
        )
        self.conn.commit()

    @staticmethod
    def _candidate_months(directory: Path) -> list[tuple[str, Path]]:
        result: list[tuple[str, Path]] = []
        if not directory.exists():
            return result
        for candidate in directory.glob("customer_location_*.parquet"):
            month = candidate.stem.replace("customer_location_", "", 1).strip()
            if (
                len(month) == 6
                and month.isdigit()
                and month.startswith("20")
            ):
                result.append((month, candidate))
        return result

    def _master_path(self, month: str) -> Path | None:
        directory = self.output_dir / "customer_location"
        requested = str(month).strip()
        exact = directory / f"customer_location_{requested}.parquet"
        if exact.exists():
            return exact

        candidates = self._candidate_months(directory)
        if not candidates:
            return None
        try:
            requested_number = int(requested)
        except ValueError:
            return None
        candidates.sort(
            key=lambda item: (
                abs(int(item[0]) - requested_number),
                int(item[0]),
            )
        )
        selected_month, path = candidates[0]
        logger.warning(
            "Exact coordinate master not found for month %s; using nearest CUSTOMER_LOCATION master month %s: %s",
            requested,
            selected_month,
            path,
        )
        return path

    def _load_month(self, month: str) -> bool:
        month = str(month).strip()
        row = self.conn.execute(
            "SELECT available FROM loaded_months WHERE month = ?",
            (month,),
        ).fetchone()
        if row is not None:
            return bool(row[0])

        path = self._master_path(month)
        if path is None:
            self.conn.execute(
                "INSERT INTO loaded_months(month, available) VALUES (?, 0)",
                (month,),
            )
            self.conn.commit()
            logger.warning("Coordinate master not found for DLPD month %s", month)
            return False

        try:
            import pyarrow.parquet as pq

            parquet = pq.ParquetFile(path)
            required = set(parquet.schema_arrow.names)
            canonical = {"IDPEL", "KOORDINAT_X", "KOORDINAT_Y"}
            if not canonical.issubset(required):
                logger.error(
                    "Coordinate master missing canonical columns: %s",
                    sorted(canonical - required),
                )
                self.conn.execute(
                    "INSERT INTO loaded_months(month, available) VALUES (?, 0)",
                    (month,),
                )
                self.conn.commit()
                return False

            inserted = 0
            for batch in parquet.iter_batches(
                batch_size=10_000,
                columns=["IDPEL", "KOORDINAT_X", "KOORDINAT_Y"],
            ):
                frame = batch.to_pandas()
                if frame.empty:
                    continue

                frame["IDPEL"] = MonthlyMerger._normalize_idpel_series(
                    frame["IDPEL"]
                )
                frame = frame.loc[frame["IDPEL"].ne("")].copy()
                if frame.empty:
                    continue

                x = pd.to_numeric(frame["KOORDINAT_X"], errors="coerce")
                y = pd.to_numeric(frame["KOORDINAT_Y"], errors="coerce")
                self.conn.executemany(
                    """
                    INSERT OR REPLACE INTO coordinate_index(
                        month, idpel, koordinat_x, koordinat_y
                    ) VALUES (?, ?, ?, ?)
                    """,
                    (
                        (
                            month,
                            idpel,
                            x_value if pd.notna(x_value) else None,
                            y_value if pd.notna(y_value) else None,
                        )
                        for idpel, x_value, y_value in zip(
                            frame["IDPEL"].tolist(),
                            x.tolist(),
                            y.tolist(),
                        )
                    ),
                )
                inserted += len(frame)
                del frame, x, y

            self.conn.execute(
                "INSERT INTO loaded_months(month, available) VALUES (?, 1)",
                (month,),
            )
            self.conn.commit()
            logger.info(
                "COORDINATE INDEX READY | month=%s | rows=%s | source=%s",
                month,
                inserted,
                path.name,
            )
            return True
        except Exception:
            logger.exception(
                "Failed to build disk-backed coordinate index | month=%s | source=%s",
                month,
                path,
            )
            self.conn.execute(
                "INSERT OR REPLACE INTO loaded_months(month, available) VALUES (?, 0)",
                (month,),
            )
            self.conn.commit()
            return False

    def enrich(self, frame: pd.DataFrame, month: str) -> pd.DataFrame:
        frame = frame.copy()
        if "IDPEL" not in frame.columns:
            frame["KOORDINAT_X"] = pd.NA
            frame["KOORDINAT_Y"] = pd.NA
            return frame

        frame["IDPEL"] = MonthlyMerger._normalize_idpel_series(frame["IDPEL"])
        frame = frame.drop(
            columns=["KOORDINAT_X", "KOORDINAT_Y", "LATITUDE", "LONGITUDE"],
            errors="ignore",
        )

        if not self._load_month(month):
            frame["KOORDINAT_X"] = pd.NA
            frame["KOORDINAT_Y"] = pd.NA
            return frame

        self.conn.execute("DELETE FROM lookup_ids")
        ids = frame["IDPEL"].drop_duplicates().tolist()
        self.conn.executemany(
            "INSERT OR IGNORE INTO lookup_ids(idpel) VALUES (?)",
            ((value,) for value in ids if value),
        )
        lookup = pd.read_sql_query(
            """
            SELECT c.idpel AS IDPEL,
                   c.koordinat_x AS KOORDINAT_X,
                   c.koordinat_y AS KOORDINAT_Y
            FROM coordinate_index AS c
            INNER JOIN lookup_ids AS l ON l.idpel = c.idpel
            WHERE c.month = ?
            """,
            self.conn,
            params=[str(month).strip()],
        )

        logger.info(
            "DLPD coordinate IDPEL diagnostic | dlpd_unique=%s | matched_unique=%s | month=%s",
            len(ids),
            len(lookup),
            month,
        )

        enriched = frame.merge(
            lookup,
            on="IDPEL",
            how="left",
            validate="many_to_one",
        )
        matched = enriched["KOORDINAT_X"].notna() & enriched["KOORDINAT_Y"].notna()
        logger.info(
            "DLPD COORDINATE ENRICHMENT | rows=%s | matched=%s | missing=%s | coverage=%.2f%% | month=%s",
            len(enriched),
            int(matched.sum()),
            int((~matched).sum()),
            (float(matched.mean()) * 100.0) if len(enriched) else 0.0,
            month,
        )
        return enriched

    def close(self) -> None:
        self.conn.close()


def _publish_staged_outputs(
    output_dir: Path,
    dataset: str,
    staged_files: list[Path],
) -> dict[str, Path]:
    """Fallback publisher; main installs the month-preserving guard."""
    folder = output_dir / "dlpd"
    folder.mkdir(parents=True, exist_ok=True)
    prepared: list[tuple[Path, Path]] = []
    for staged in staged_files:
        final = folder / staged.name
        hidden = folder / f".{staged.name}.new"
        os.replace(staged, hidden)
        prepared.append((hidden, final))

    published: dict[str, Path] = {}
    try:
        for hidden, final in prepared:
            os.replace(hidden, final)
            parts = final.stem.split("_")
            month = (
                parts[-2]
                if len(parts) >= 3 and parts[-1].startswith("part")
                else ""
            )
            if month:
                published.setdefault(month, final)
    except Exception:
        for hidden, _ in prepared:
            hidden.unlink(missing_ok=True)
        raise

    if not published:
        raise RuntimeError(
            "DLPD publish completed but no monthly partitions were detected."
        )
    return published


def _process_all_dlpds(
    dataset: str,
    files: list[Path],
    output_dir: Path,
) -> dict[str, Path]:
    """Scan all DLPD sources once and publish monthly parquet partitions."""
    if not files:
        raise ValueError(f"No files supplied for dataset '{dataset}'.")

    output_dir = Path(output_dir)
    staging_root = output_dir / ".dlpd_stream_staging"
    staging_dir = staging_root / f"{dataset.lower()}_{uuid.uuid4().hex}"
    staging_dir.mkdir(parents=True, exist_ok=True)

    prefix = (
        "dlpd_pascabayar_"
        if dataset == "DLPD_PASCABAYAR"
        else "dlpd_prabayar_"
    )
    part_numbers: dict[str, int] = {}
    staged_files: list[Path] = []

    seen_db = sqlite3.connect(staging_dir / "seen_idpel.sqlite3")
    coordinate_index = _CoordinateIndex(
        staging_dir / "coordinate_index.sqlite3",
        output_dir,
    )

    try:
        seen_db.execute("CREATE TABLE seen_idpel (idpel TEXT PRIMARY KEY)")
        seen_db.execute("CREATE TEMP TABLE batch_ids (idpel TEXT PRIMARY KEY)")
        seen_db.commit()

        for source in files:
            source = Path(source)
            logger.info(
                "STREAMING DLPD SOURCE | dataset=%s | file=%s | size=%s | chunk_rows=%s",
                dataset,
                source.name,
                source.stat().st_size,
                _CHUNK_ROWS,
            )

            for chunk_number, chunk in enumerate(
                _iter_excel_chunks(source, dataset),
                start=1,
            ):
                if chunk.empty:
                    continue

                chunk = MonthlyMerger._normalize_dataframe_columns(chunk)
                chunk["SOURCE_FILE"] = source.name
                chunk = _deduplicate_first_idpel(chunk, seen_db)
                if chunk.empty:
                    continue

                transformed = DLPDTransformer().transform(chunk)
                if transformed.empty or "MONTH" not in transformed.columns:
                    continue

                transformed["MONTH"] = (
                    transformed["MONTH"]
                    .apply(DLPDTransformer._normalize_month_value)
                    .fillna("")
                    .astype(str)
                    .str.strip()
                )

                for month, frame in transformed.groupby(
                    "MONTH",
                    sort=True,
                    dropna=False,
                ):
                    month = str(month).strip()
                    if not month:
                        continue

                    frame = coordinate_index.enrich(frame, month)
                    frame = _normalise_output_frame(frame)

                    part_numbers[month] = part_numbers.get(month, 0) + 1
                    output = (
                        staging_dir
                        / f"{prefix}{month}_part{part_numbers[month]:05d}.parquet"
                    )
                    frame.to_parquet(output, index=False)
                    staged_files.append(output)

                    logger.info(
                        "STREAMING DLPD PART WRITTEN | dataset=%s | month=%s | chunk=%s | rows=%s",
                        dataset,
                        month,
                        chunk_number,
                        len(frame),
                    )

        if not staged_files:
            raise ValueError(
                "Streaming DLPD processing produced no monthly rows "
                f"for {dataset}."
            )

        outputs = _publish_staged_outputs(
            output_dir=output_dir,
            dataset=dataset,
            staged_files=staged_files,
        )
        logger.info(
            "STREAMING DLPD COMPLETE | dataset=%s | months=%s | parts=%s",
            dataset,
            sorted(outputs),
            len(staged_files),
        )
        return outputs
    finally:
        coordinate_index.close()
        seen_db.close()
        shutil.rmtree(staging_dir, ignore_errors=True)
        try:
            staging_root.rmdir()
        except OSError:
            pass


def _streaming_merge(
    dataset: str,
    month: str | None,
    files: list[Path],
    output_dir: Path,
) -> Path:
    if not files:
        raise ValueError(f"No files supplied for dataset '{dataset}'.")

    key = _source_key(dataset, files)
    target = (
        DLPDTransformer._normalize_month_value(month)
        if month is not None
        else None
    )
    outputs = _COMPLETED_RUNS.get(key)
    if outputs is None:
        outputs = _process_all_dlpds(
            dataset=dataset,
            files=files,
            output_dir=output_dir,
        )
        _COMPLETED_RUNS[key] = outputs

    if target:
        output = outputs.get(target)
        if output is None:
            raise ValueError(
                f"DLPD month {target} was requested but no rows were found."
            )
        return output
    return next(iter(outputs.values()))


def install_streaming_dlpd_merger_patch() -> None:
    global _INSTALLED
    if _INSTALLED:
        return

    def patched_merge(
        dataset: str,
        month: str | None,
        files: list[Path],
        output_dir: Path,
    ) -> Path:
        if dataset in MonthlyMerger.COORDINATE_DATASETS:
            return _streaming_merge(dataset, month, files, output_dir)
        return _ORIGINAL_MERGE(dataset, month, files, output_dir)

    MonthlyMerger.merge = staticmethod(patched_merge)
    _INSTALLED = True
    logger.info(
        "Installed restart-safe memory-bounded DLPD merger patch | chunk_rows=%s",
        _CHUNK_ROWS,
    )
