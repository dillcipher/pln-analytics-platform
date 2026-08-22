from __future__ import annotations

from pathlib import Path


class FileDetector:
    """
    Detect dataset type from uploaded filename and, when the filename is
    generic/renamed, from the Excel header schema.

    Filename detection remains the fast path. Content detection is only a
    lightweight header inspection using openpyxl read-only mode; it never
    loads the workbook into memory.
    """

    ANEV = "ANEV"
    DLPD_PASCABAYAR = "DLPD_PASCABAYAR"
    DLPD_PRABAYAR = "DLPD_PRABAYAR"
    PENGECEKAN = "PENGECEKAN"
    CUSTOMER_LOCATION = "CUSTOMER_LOCATION"
    UNKNOWN = "UNKNOWN"

    COORDINATE_MASTER_FILENAMES = {
        "to_prabayar.xlsx",
        "to_pascabayar.xlsx",
    }

    @staticmethod
    def _normalize_filename(filepath: Path) -> str:
        name = Path(filepath).name.lower().strip()
        name = name.replace("-", "_").replace(" ", "_")
        while "__" in name:
            name = name.replace("__", "_")
        return name

    @staticmethod
    def _normalize_column(value: object) -> str:
        text = str(value or "").replace("\n", " ").strip().upper()
        return "".join(char for char in text if char.isalnum())

    @classmethod
    def is_coordinate_master(cls, filepath: Path) -> bool:
        return cls._normalize_filename(filepath) in cls.COORDINATE_MASTER_FILENAMES

    @classmethod
    def _detect_from_excel_schema(cls, filepath: Path) -> str:
        """
        Detect a renamed/generic Excel upload from header columns.

        The scan is intentionally bounded. It checks several worksheets and
        the first 50 rows so files with title/metadata rows before the real
        header are still detected, while openpyxl read-only mode keeps memory
        usage low for large workbooks.
        """
        path = Path(filepath)
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return cls.UNKNOWN
        if not path.exists() or not path.is_file():
            return cls.UNKNOWN

        try:
            from openpyxl import load_workbook

            workbook = load_workbook(
                path,
                read_only=True,
                data_only=True,
            )

            try:
                for worksheet in workbook.worksheets[:8]:
                    for row in worksheet.iter_rows(
                        min_row=1,
                        max_row=50,
                        values_only=True,
                    ):
                        columns = {
                            cls._normalize_column(value)
                            for value in row
                            if value is not None
                        }

                        # Most specific signatures first. A DLPD workbook
                        # also commonly contains IDPEL, so DLPD must win over
                        # the generic PENGECEKAN IDPEL+NAMA signature.
                        if {"IDPEL", "THBLREK", "DLPD"}.issubset(columns):
                            return cls.DLPD_PASCABAYAR

                        if {"IDPEL", "THBL", "DLPD"}.issubset(columns):
                            return cls.DLPD_PRABAYAR

                        if "IDPEL" in columns and (
                            {"LATITUDE", "LONGITUDE"}.issubset(columns)
                            or {"KOORDINATX", "KOORDINATY"}.issubset(columns)
                        ):
                            return cls.CUSTOMER_LOCATION

                        if {"LOCATIONCODE", "READDATE", "SUSPECTNAME"}.issubset(columns):
                            return cls.ANEV

                        if {"IDPEL", "NAMA"}.issubset(columns):
                            return cls.PENGECEKAN
            finally:
                workbook.close()

        except Exception:
            # Detection must remain non-fatal. DatasetValidator/ETL will
            # provide the detailed workbook error once a dataset is known.
            return cls.UNKNOWN

        return cls.UNKNOWN

    @classmethod
    def detect(cls, filepath: Path) -> str:
        """
        Detect the dataset represented by a file.

        Detection order:
            1. Exact coordinate masters
            2. Filename rules
            3. Excel schema fallback for generic/renamed files
            4. UNKNOWN
        """
        path = Path(filepath)
        name = cls._normalize_filename(path)

        if cls.is_coordinate_master(path):
            return cls.CUSTOMER_LOCATION

        if "dil_saldo_mask" in name or "dil_saldo" in name:
            return cls.CUSTOMER_LOCATION

        if (
            "dlpd" in name
            and (
                "pascabayar" in name
                or "pasca_bayar" in name
                or ("pln" in name and "prabayar" not in name)
            )
        ):
            return cls.DLPD_PASCABAYAR

        if (
            "dlpd" in name
            and (
                "tidak_beli_token" in name
                or "prabayar" in name
                or "pra_bayar" in name
            )
        ):
            return cls.DLPD_PRABAYAR

        if "pengecekan" in name:
            return cls.PENGECEKAN

        if "anev" in name or "17_anev" in name:
            return cls.ANEV

        return cls._detect_from_excel_schema(path)
