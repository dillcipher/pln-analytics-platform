"""Memory-safe DLPD month resolution for large XLSX workbooks.

The normal MonthResolver implementation can materialize an entire workbook.
This patch keeps the existing month business rules but makes the streaming
pass substantially cheaper: only the required columns are iterated instead of
allocating every column of every row, and the workbook is always closed.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.etl.detector.month_resolver import MonthResolver
from app.etl.validator.validator import DatasetValidator


_INSTALLED = False


def _stream_read_dlpd_months(
    cls,
    filepath: Path,
    dataset: str,
    sheet_name: str,
) -> list[str]:
    """Resolve all DLPD business months without materializing the workbook."""
    del dataset  # The DLPD month rules are shared by both customer types.
    filepath = Path(filepath)

    thbl_column = DatasetValidator.normalize_column("THBL")
    thblrek_column = DatasetValidator.normalize_column("THBLREK")
    date_column = DatasetValidator.normalize_column("DLPD_TGLBACA")
    required = {thbl_column, thblrek_column, date_column}

    workbook = None
    try:
        workbook = load_workbook(
            filename=filepath,
            read_only=True,
            data_only=True,
        )

        if not workbook.sheetnames:
            return []

        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            normalized_sheets = {
                str(name).strip().upper(): name
                for name in workbook.sheetnames
            }
            worksheet = workbook[
                normalized_sheets.get(
                    str(sheet_name).strip().upper(),
                    workbook.sheetnames[0],
                )
            ]

        header_row: int | None = None
        header_indexes: dict[str, int] = {}

        # Preserve the existing detector behavior: a DLPD header is normally
        # identified by IDPEL/LOCATIONCODE within the first 15 rows.
        for row_index, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=15,
                values_only=True,
            ),
            start=1,
        ):
            normalized_row = [
                DatasetValidator.normalize_column(value)
                for value in row
            ]
            indexes: dict[str, int] = {}
            for column_index, normalized in enumerate(normalized_row):
                if normalized in required and normalized not in indexes:
                    indexes[normalized] = column_index

            if "IDPEL" in normalized_row or "LOCATIONCODE" in normalized_row:
                header_row = row_index
                header_indexes = indexes
                break

        if header_row is None:
            header_row = 1
            header_values = next(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=1,
                    values_only=True,
                ),
                (),
            )
            header_indexes = {
                DatasetValidator.normalize_column(value): index
                for index, value in enumerate(header_values)
                if DatasetValidator.normalize_column(value) in required
            }

        if not header_indexes:
            return []

        # openpyxl otherwise creates a tuple for every cell in every row.
        # Restrict iteration to the smallest span containing the three fields
        # so a wide DLPD workbook does not temporarily allocate hundreds of
        # unrelated cells per row.
        absolute_indexes = sorted(header_indexes.values())
        min_col = absolute_indexes[0] + 1
        max_col = absolute_indexes[-1] + 1
        relative_indexes = {
            key: index - absolute_indexes[0]
            for key, index in header_indexes.items()
        }

        months: set[str] = set()
        rows_seen = 0

        for row in worksheet.iter_rows(
            min_row=header_row + 1,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            rows_seen += 1

            thbl_index = relative_indexes.get(thbl_column)
            thblrek_index = relative_indexes.get(thblrek_column)
            date_index = relative_indexes.get(date_column)

            thbl = (
                row[thbl_index]
                if thbl_index is not None and thbl_index < len(row)
                else None
            )
            thblrek = (
                row[thblrek_index]
                if thblrek_index is not None and thblrek_index < len(row)
                else None
            )
            detail_date = (
                row[date_index]
                if date_index is not None and date_index < len(row)
                else None
            )

            thbl_start = cls._month_start(thbl)
            thblrek_start = cls._month_start(thblrek)
            parsed_date = cls._parse_date(detail_date)

            if (
                parsed_date is not None
                and thbl_start is not None
                and thblrek_start is not None
            ):
                period_start = min(thbl_start, thblrek_start)
                period_month = max(thbl_start, thblrek_start)
                period_end = cls._month_end(period_month)

                if (
                    period_end is not None
                    and period_start <= parsed_date <= period_end
                ):
                    months.add(parsed_date.strftime("%Y%m"))
                    continue

            if (
                parsed_date is not None
                and (thbl_start is not None or thblrek_start is not None)
            ):
                months.add(parsed_date.strftime("%Y%m"))
                continue

            month = cls._normalize_month(thblrek)
            if month:
                months.add(month)
                continue

            month = cls._normalize_month(thbl)
            if month:
                months.add(month)

        return sorted(months)

    except Exception as exc:
        print(
            "[MonthResolver][streaming] "
            f"{filepath.name}: {exc}"
        )
        return []
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass


def install_streaming_month_resolver_patch() -> None:
    """Install the memory-safe DLPD reader once per process."""
    global _INSTALLED
    if _INSTALLED:
        return

    MonthResolver._read_dlpd_months = classmethod(_stream_read_dlpd_months)
    _INSTALLED = True
