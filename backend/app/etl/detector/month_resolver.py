from __future__ import annotations

import re
from pathlib import Path

import pandas as pd

from app.etl.detector.detector import FileDetector
from app.etl.validator.validator import DatasetValidator


class MonthResolver:
    """Resolve DLPD business months without materialising huge workbooks."""

    MONTH_PATTERN = re.compile(r"^(20\d{2})(0[1-9]|1[0-2])$")
    MONTH_SEARCH_PATTERN = re.compile(r"(20\d{2})(0[1-9]|1[0-2])")
    COORDINATE_MASTER_FILES = {"to_prabayar.xlsx", "to_pascabayar.xlsx"}

    @staticmethod
    def _normalized_filename(filepath: Path) -> str:
        name = Path(filepath).name.lower().strip().replace("-", "_").replace(" ", "_")
        while "__" in name:
            name = name.replace("__", "_")
        return name

    @classmethod
    def is_coordinate_master(cls, filepath: Path) -> bool:
        return FileDetector.is_coordinate_master(filepath) or cls._normalized_filename(filepath) in cls.COORDINATE_MASTER_FILES

    @classmethod
    def _normalize_month(cls, value) -> str | None:
        if value is None:
            return None
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass
        if isinstance(value, pd.Timestamp):
            try:
                return value.strftime("%Y%m")
            except Exception:
                return None
        if hasattr(value, "year") and hasattr(value, "month"):
            try:
                candidate = f"{int(value.year):04d}{int(value.month):02d}"
                if cls.MONTH_PATTERN.fullmatch(candidate):
                    return candidate
            except Exception:
                pass
        if isinstance(value, (int, float)):
            try:
                text = str(int(value))
                if cls.MONTH_PATTERN.fullmatch(text):
                    return text
            except Exception:
                pass
        text = str(value).strip()
        if not text:
            return None
        if cls.MONTH_PATTERN.fullmatch(text):
            return text
        match = cls.MONTH_SEARCH_PATTERN.search(text)
        if match:
            return f"{match.group(1)}{match.group(2)}"
        try:
            parsed = pd.to_datetime(text, errors="coerce")
            if not pd.isna(parsed):
                candidate = parsed.strftime("%Y%m")
                if cls.MONTH_PATTERN.fullmatch(candidate):
                    return candidate
        except Exception:
            pass
        return None

    @classmethod
    def _parse_date(cls, value) -> pd.Timestamp | None:
        if value is None:
            return None
        if isinstance(value, pd.Timestamp):
            return value
        if hasattr(value, "year") and hasattr(value, "month"):
            try:
                return pd.Timestamp(value)
            except Exception:
                return None
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass
        try:
            parsed = pd.to_datetime(value, errors="coerce")
            return None if pd.isna(parsed) else pd.Timestamp(parsed)
        except Exception:
            return None

    @classmethod
    def _month_start(cls, value) -> pd.Timestamp | None:
        month = cls._normalize_month(value)
        if not month:
            return None
        try:
            return pd.Timestamp(year=int(month[:4]), month=int(month[4:6]), day=1)
        except Exception:
            return None

    @staticmethod
    def _month_end(value: pd.Timestamp | None) -> pd.Timestamp | None:
        if value is None:
            return None
        try:
            return value + pd.offsets.MonthEnd(1) + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1)
        except Exception:
            return None

    @classmethod
    def _get_sheet_name(cls, filepath: Path, dataset: str) -> str:
        return DatasetValidator.get_sheet_name(filepath, dataset)

    @classmethod
    def _read_dlpd_months_streaming(cls, filepath: Path, dataset: str, sheet_name: str, header: int) -> list[str]:
        """Scan only THBL/THBLREK/DLPD_TGLBACA using openpyxl read-only mode."""
        from openpyxl import load_workbook

        wb = load_workbook(filepath, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]
            header_values = next(ws.iter_rows(min_row=header + 1, max_row=header + 1, values_only=True), None)
            if not header_values:
                return []

            normalized = [DatasetValidator.normalize_column(v) for v in header_values]
            thbl_key = DatasetValidator.normalize_column("THBL")
            thblrek_key = DatasetValidator.normalize_column("THBLREK")
            date_key = DatasetValidator.normalize_column("DLPD_TGLBACA")
            indexes = {name: i for i, name in enumerate(normalized) if name in {thbl_key, thblrek_key, date_key}}
            thbl_idx = indexes.get(thbl_key)
            thblrek_idx = indexes.get(thblrek_key)
            date_idx = indexes.get(date_key)

            months: set[str] = set()
            for row in ws.iter_rows(min_row=header + 2, values_only=True):
                thbl = row[thbl_idx] if thbl_idx is not None and thbl_idx < len(row) else None
                thblrek = row[thblrek_idx] if thblrek_idx is not None and thblrek_idx < len(row) else None
                detail_date = row[date_idx] if date_idx is not None and date_idx < len(row) else None

                thbl_start = cls._month_start(thbl)
                thblrek_start = cls._month_start(thblrek)
                parsed_date = cls._parse_date(detail_date)

                if parsed_date is not None and thbl_start is not None and thblrek_start is not None:
                    period_start = min(thbl_start, thblrek_start)
                    period_end = cls._month_end(max(thbl_start, thblrek_start))
                    if period_end is not None and period_start <= parsed_date <= period_end:
                        months.add(parsed_date.strftime("%Y%m"))
                        continue

                if parsed_date is not None and (thbl_start is not None or thblrek_start is not None):
                    months.add(parsed_date.strftime("%Y%m"))
                    continue

                month = cls._normalize_month(thblrek) or cls._normalize_month(thbl)
                if month:
                    months.add(month)

            return sorted(months)
        finally:
            wb.close()

    @classmethod
    def _read_dlpd_months(cls, filepath: Path, dataset: str, sheet_name: str) -> list[str]:
        try:
            header = DatasetValidator.detect_header_row(filepath=filepath, sheet_name=sheet_name)
            if Path(filepath).stat().st_size >= 20 * 1024 * 1024:
                return cls._read_dlpd_months_streaming(filepath, dataset, sheet_name, header)

            dataframe = pd.read_excel(
                filepath,
                sheet_name=sheet_name,
                header=header,
                usecols=lambda column: DatasetValidator.normalize_column(column) in {
                    DatasetValidator.normalize_column("THBL"),
                    DatasetValidator.normalize_column("THBLREK"),
                    DatasetValidator.normalize_column("DLPD_TGLBACA"),
                },
            )
            if dataframe.empty:
                return []
            dataframe.columns = [DatasetValidator.normalize_column(c) for c in dataframe.columns]
            thbl_column = DatasetValidator.normalize_column("THBL")
            thblrek_column = DatasetValidator.normalize_column("THBLREK")
            date_column = DatasetValidator.normalize_column("DLPD_TGLBACA")
            for column in (thbl_column, thblrek_column, date_column):
                if column not in dataframe.columns:
                    dataframe[column] = None

            months: set[str] = set()
            for _, row in dataframe.iterrows():
                thbl = row.get(thbl_column)
                thblrek = row.get(thblrek_column)
                detail_date = row.get(date_column)
                thbl_start = cls._month_start(thbl)
                thblrek_start = cls._month_start(thblrek)
                parsed_date = cls._parse_date(detail_date)
                if parsed_date is not None and thbl_start is not None and thblrek_start is not None:
                    period_start = min(thbl_start, thblrek_start)
                    period_end = cls._month_end(max(thbl_start, thblrek_start))
                    if period_end is not None and period_start <= parsed_date <= period_end:
                        months.add(parsed_date.strftime("%Y%m"))
                        continue
                if parsed_date is not None and (thbl_start is not None or thblrek_start is not None):
                    months.add(parsed_date.strftime("%Y%m"))
                    continue
                month = cls._normalize_month(thblrek) or cls._normalize_month(thbl)
                if month:
                    months.add(month)
            return sorted(months)
        except Exception as exc:
            print(f"[MonthResolver] {filepath.name}: {exc}")
            return []

    @classmethod
    def _read_first_month(cls, filepath: Path, dataset: str, sheet_name: str, column_name: str) -> str | None:
        try:
            header = DatasetValidator.detect_header_row(filepath=filepath, sheet_name=sheet_name)
            dataframe = pd.read_excel(filepath, sheet_name=sheet_name, header=header, usecols=lambda column: DatasetValidator.normalize_column(column) == DatasetValidator.normalize_column(column_name))
            if dataframe.empty:
                return None
            for value in dataframe.iloc[:, 0].tolist():
                month = cls._normalize_month(value)
                if month:
                    return month
            return None
        except Exception as exc:
            print(f"[MonthResolver] {filepath.name}: {exc}")
            return None

    @classmethod
    def resolve_months(cls, filepath: Path, dataset: str) -> list[str]:
        filepath = Path(filepath)
        if cls.is_coordinate_master(filepath):
            return []
        sheet_name = cls._get_sheet_name(filepath, dataset)
        if dataset in {"DLPD_PASCABAYAR", "DLPD_PRABAYAR"}:
            return cls._read_dlpd_months(filepath, dataset, sheet_name)
        column_name = {"ANEV": "READ_DATE", "PENGECEKAN": "WAKTU_PERIKSA", "CUSTOMER_LOCATION": "MONTH"}.get(dataset, "MONTH")
        month = cls._read_first_month(filepath, dataset, sheet_name, column_name)
        return [month] if month else []
